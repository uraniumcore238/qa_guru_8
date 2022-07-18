import csv
import sys

from openpyxl import load_workbook
import zipfile
import PyPDF2



def read_the_files(file_end):
    archive = 'my_test_zip.zip'
    zip_file = zipfile.ZipFile(archive)

    for file in zip_file.infolist():
        if file.filename.endswith('.pdf') and file_end == '.pdf':
            print('PDF')
            pdf_file = zip_file.extract(file.filename, 'tmp')
            plread = PyPDF2.PdfFileReader(pdf_file)
            getpage37 = plread.getPage(37)
            text37 = getpage37.extractText()
            assert '2.3.9Factories as fixtures' in text37
            sys.exit()
        if file.filename.endswith('.csv') and file_end == '.csv':
            print('CSV')
            csv_file = zip_file.extract(file.filename, 'tmp')
            with open(csv_file) as csvfile:
                csvfile = csv.reader(csvfile)
                for strings in csvfile:
                    if 'Dollars (millions)' in strings:
                        assert True
                        break

                else:
                    assert False
            sys.exit()
        if file.filename.endswith('.xlsx') and file_end == '.xlsx':
            print('XLSX')
            xlsx_file = zip_file.extract(file.filename, 'tmp')
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            list_fom_cells = []
            for row in range(sheet.max_row):
                for column in range(sheet.max_column):
                    value_str = str(sheet.cell(row=(row+1), column=(column+1)).value)
                    list_fom_cells.append(value_str)
            if '1' in list_fom_cells:
                assert True

            else:
                assert False
            sys.exit()


read_the_files('.csv')
